"""Template ingestion validation tests."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from simple_e2e_tester.configuration.runtime_settings import SchemaConfig
from simple_e2e_tester.schema_management import flatten_schema, load_schema_document
from simple_e2e_tester.template_generation import TEMPLATE_SHEET_NAME, generate_template_workbook
from simple_e2e_tester.template_ingestion.testcase_models import TemplateTestCase
from simple_e2e_tester.template_ingestion.workbook_reader import (
    TemplateValidationError,
    read_template,
)


def _schema_config() -> SchemaConfig:
    schema_text = json.dumps(
        {
            "type": "object",
            "properties": {
                "emailabsender": {"type": "string"},
                "emailbetreff": {"type": "string"},
                "ki_ergebnis": {
                    "type": "object",
                    "properties": {
                        "klasse": {"type": "array", "items": {"type": "string"}},
                        "fachdaten": {
                            "type": "object",
                            "properties": {
                                "grund": {
                                    "type": "object",
                                    "properties": {
                                        "value": {"type": "string"},
                                        "score": {"type": "number"},
                                    },
                                }
                            },
                        },
                    },
                },
            },
        }
    )
    return SchemaConfig(schema_type="json_schema", text=schema_text, source_path=None)


def _write_template(tmp_path: Path) -> tuple[Path, list[str]]:
    schema_config = _schema_config()
    fields = flatten_schema(load_schema_document(schema_config))
    output_path = tmp_path / "template.xlsx"
    generate_template_workbook(schema_config, fields, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }

    def set_row(row: int, values: dict[str, object]) -> None:
        for column_name, value in values.items():
            sheet.cell(row=row, column=header_map[column_name], value=cast_any(value))

    set_row(
        3,
        {
            "ID": "TC-1",
            "Tags": " smoke ,happy ",
            "Enabled": None,
            "Notes": "baseline",
            "FROM": "sender@example.com",
            "SUBJECT": "Test 1",
            "BODY": "text body",
            "ATTACHMENT": "",
            "emailabsender": "expected-from",
            "emailbetreff": "expected-subject",
            "ki_ergebnis.fachdaten.grund.value": "hello",
        },
    )
    set_row(
        4,
        {
            "ID": "TC-2",
            "Tags": "",
            "Enabled": "FALSE",
            "Notes": "skip me",
            "FROM": "second@example.com",
            "SUBJECT": "Test 2",
            "BODY": "",
            "ATTACHMENT": "",
        },
    )

    workbook.save(output_path)
    return output_path, [field.path for field in fields]


def test_read_template_parses_rows(tmp_path: Path) -> None:
    template_path, field_names = _write_template(tmp_path)

    result = read_template(template_path, field_names)

    assert len(result.testcases) == 2
    first = result.testcases[0]
    assert isinstance(first, TemplateTestCase)
    assert first.test_id == "TC-1"
    assert first.tags == ("smoke", "happy")
    assert first.enabled is True
    assert first.from_address == "sender@example.com"
    assert first.subject == "Test 1"
    assert first.expected_values["emailabsender"] == "expected-from"
    assert first.expected_values["ki_ergebnis.fachdaten.grund.value"] == "hello"
    second = result.testcases[1]
    assert second.enabled is False


def test_read_template_validates_expected_columns(tmp_path: Path) -> None:
    template_path, field_names = _write_template(tmp_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    sheet.cell(row=2, column=sheet.max_column).value = "unexpected"
    workbook.save(template_path)

    with pytest.raises(TemplateValidationError):
        read_template(template_path, field_names)


def test_read_template_detects_duplicate_ids(tmp_path: Path) -> None:
    template_path, field_names = _write_template(tmp_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    sheet.cell(row=4, column=_column_for(sheet, "ID")).value = "TC-1"
    workbook.save(template_path)

    with pytest.raises(TemplateValidationError):
        read_template(template_path, field_names)


def test_read_template_detects_duplicate_from_subject(tmp_path: Path) -> None:
    template_path, field_names = _write_template(tmp_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    sheet.cell(row=4, column=_column_for(sheet, "ID")).value = "TC-3"
    sheet.cell(row=4, column=_column_for(sheet, "FROM")).value = "sender@example.com"
    sheet.cell(row=4, column=_column_for(sheet, "SUBJECT")).value = "Test 1"
    sheet.cell(row=4, column=_column_for(sheet, "Enabled")).value = True
    workbook.save(template_path)

    with pytest.raises(TemplateValidationError):
        read_template(template_path, field_names)


def test_read_template_validates_email_format(tmp_path: Path) -> None:
    template_path, field_names = _write_template(tmp_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    sheet.cell(row=3, column=_column_for(sheet, "FROM")).value = "invalid-email"
    workbook.save(template_path)

    with pytest.raises(TemplateValidationError):
        read_template(template_path, field_names)


def test_read_template_errors_when_expected_fields_are_empty(tmp_path: Path) -> None:
    template_path, _ = _write_template(tmp_path)

    with pytest.raises(TemplateValidationError, match="Expected fields list must not be empty"):
        read_template(template_path, [])


def test_read_template_errors_on_invalid_group_headers(tmp_path: Path) -> None:
    template_path, field_names = _write_template(tmp_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    sheet.cell(row=1, column=1).value = "Meta"
    workbook.save(template_path)

    with pytest.raises(TemplateValidationError, match="missing required group headers"):
        read_template(template_path, field_names)


def test_read_template_errors_on_unparseable_enabled_value(tmp_path: Path) -> None:
    template_path, field_names = _write_template(tmp_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    sheet.cell(row=3, column=_column_for(sheet, "Enabled")).value = "MAYBE"
    workbook.save(template_path)

    with pytest.raises(TemplateValidationError, match="Unable to interpret boolean value"):
        read_template(template_path, field_names)


def test_read_template_errors_when_no_testcase_rows_exist(tmp_path: Path) -> None:
    template_path, field_names = _write_template(tmp_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    if sheet.max_row >= 3:
        sheet.delete_rows(3, amount=sheet.max_row - 2)
    workbook.save(template_path)

    with pytest.raises(TemplateValidationError, match="does not contain any testcase rows"):
        read_template(template_path, field_names)


def _column_for(sheet, header: str) -> int:
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=2, column=col).value == header:
            return col
    raise AssertionError(f"Column {header} not found")


def cast_any(value: object) -> Any:
    return value
