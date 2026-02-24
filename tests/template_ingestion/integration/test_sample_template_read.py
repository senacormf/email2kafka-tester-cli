"""Integration test for reading sample schema template."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from simple_e2e_tester.configuration.runtime_settings import SchemaConfig
from simple_e2e_tester.schema_management import flatten_schema, load_schema_document
from simple_e2e_tester.template_generation import TEMPLATE_SHEET_NAME, generate_template_workbook
from simple_e2e_tester.template_ingestion.workbook_reader import read_template


def test_read_sample_schema_template(tmp_path: Path) -> None:
    schema_text = (
        Path(__file__)
        .resolve()
        .parents[3]
        .joinpath("samples", "sample-json-schema.json")
        .read_text(encoding="utf-8")
    )
    schema_config = SchemaConfig(schema_type="json_schema", text=schema_text, source_path=None)
    fields = flatten_schema(load_schema_document(schema_config))

    output_path = tmp_path / "sample-template.xlsx"
    generate_template_workbook(schema_config, fields, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "SAMPLE-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sample@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Public Flow"

    workbook.save(output_path)

    result = read_template(output_path, [field.path for field in fields])

    assert len(result.testcases) == 1
    testcase = result.testcases[0]
    assert testcase.expected_values["event_id"] is None
    assert testcase.expected_values["message_subject"] is None
    assert testcase.test_id == "SAMPLE-1"
