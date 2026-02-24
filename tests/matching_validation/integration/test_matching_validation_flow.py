"""Integration tests for matching/validation with template ingestion."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from simple_e2e_tester.configuration.runtime_settings import MatchingConfig, SchemaConfig
from simple_e2e_tester.kafka_consumption.actual_event_messages import ActualEventMessage
from simple_e2e_tester.matching_validation.case_evaluator import match_and_validate
from simple_e2e_tester.matching_validation.event_boundary_mappers import (
    to_actual_events,
    to_expected_events,
)
from simple_e2e_tester.schema_management import flatten_schema, load_schema_document
from simple_e2e_tester.template_generation import TEMPLATE_SHEET_NAME, generate_template_workbook
from simple_e2e_tester.template_ingestion.workbook_reader import read_template


def test_matching_and_validation_with_generated_template(tmp_path: Path) -> None:
    schema_text = (
        Path(__file__)
        .resolve()
        .parents[3]
        .joinpath("samples", "sample-json-schema.json")
        .read_text(encoding="utf-8")
    )
    schema_config = SchemaConfig(schema_type="json_schema", text=schema_text, source_path=None)
    fields = flatten_schema(load_schema_document(schema_config))

    template_path = tmp_path / "template.xlsx"
    generate_template_workbook(schema_config, fields, template_path)

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "FLOW-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Case-A"
    sheet.cell(row=3, column=header_map["sender_address"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["message_subject"]).value = "Case-A"
    sheet.cell(row=3, column=header_map["model_output.attributes.reason.score"]).value = "1,50+-0,1"
    workbook.save(template_path)

    testcases = read_template(template_path, [field.path for field in fields]).testcases
    message = ActualEventMessage(
        key=None,
        value={},
        timestamp=datetime.now(UTC),
        flattened={
            "sender_address": "sender@example.com",
            "message_subject": "Case-A",
            "model_output.attributes.reason.score": 1.58,
        },
    )

    result = match_and_validate(
        to_expected_events(testcases),
        to_actual_events([message]),
        MatchingConfig(from_field="sender_address", subject_field="message_subject"),
        fields,
    )

    assert len(result.matches) == 1
    assert result.matches[0].expected_event.expected_event_id == "FLOW-1"
    assert result.matches[0].mismatches == ()
    assert result.conflicts == ()
    assert result.unmatched_actual_events == ()
    assert result.unmatched_expected_event_ids == ()
