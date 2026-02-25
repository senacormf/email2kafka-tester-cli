"""Scenario-style integration tests for core run behaviors."""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
from datetime import UTC, datetime
from pathlib import Path

import pytest
from click.testing import CliRunner
from openpyxl import load_workbook
from simple_e2e_tester.cli import cli
from simple_e2e_tester.configuration.loader import load_configuration
from simple_e2e_tester.configuration.runtime_settings import (
    KafkaSettings,
    MatchingConfig,
    SchemaConfig,
)
from simple_e2e_tester.email_sending.delivery_outcomes import EmailSendResult, SendStatus
from simple_e2e_tester.kafka_consumption.actual_event_messages import ActualEventMessage
from simple_e2e_tester.kafka_consumption.actual_event_reader import ActualEventReader
from simple_e2e_tester.matching_validation.case_evaluator import match_and_validate
from simple_e2e_tester.matching_validation.event_boundary_mappers import (
    to_actual_events,
    to_expected_events,
)
from simple_e2e_tester.matching_validation.matching_outcomes import MatchValidationResult
from simple_e2e_tester.results_writing import RunMetadata, write_results_workbook
from simple_e2e_tester.run_execution.run_contracts import RunRequest
from simple_e2e_tester.run_execution.validation_run_use_case import (
    execute_email_kafka_validation_run,
)
from simple_e2e_tester.schema_management import flatten_schema, load_schema_document
from simple_e2e_tester.template_generation import TEMPLATE_SHEET_NAME, generate_template_workbook
from simple_e2e_tester.template_ingestion.workbook_reader import read_template


def _schema_config() -> SchemaConfig:
    return SchemaConfig(
        schema_type="json_schema",
        text=json.dumps(
            {
                "type": "object",
                "properties": {
                    "sender": {"type": "string"},
                    "subject": {"type": "string"},
                    "score": {"type": "number"},
                },
            }
        ),
        source_path=None,
    )


def _kafka_message(*, sender: str, subject: str, score: float) -> ActualEventMessage:
    flattened = {
        "sender": sender,
        "subject": subject,
        "score": score,
    }
    return ActualEventMessage(
        key=None,
        value=flattened,
        timestamp=datetime.now(UTC),
        flattened=flattened,
    )


def _write_case_sheet(
    tmp_path: Path,
    *,
    first_subject: str,
    second_subject: str | None = None,
    expected_score: str = "1,50+-0,1",
    first_enabled: bool = True,
    second_enabled: bool = True,
) -> tuple[Path, list]:
    schema_config = _schema_config()
    fields = flatten_schema(load_schema_document(schema_config))
    template_path = tmp_path / "cases.xlsx"
    generate_template_workbook(schema_config, fields, template_path)

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }

    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = first_subject
    sheet.cell(row=3, column=header_map["score"]).value = expected_score
    sheet.cell(row=3, column=header_map["Enabled"]).value = first_enabled

    if second_subject is not None:
        sheet.cell(row=4, column=header_map["ID"]).value = "TC-2"
        sheet.cell(row=4, column=header_map["FROM"]).value = "sender@example.com"
        sheet.cell(row=4, column=header_map["SUBJECT"]).value = second_subject
        sheet.cell(row=4, column=header_map["score"]).value = expected_score
        sheet.cell(row=4, column=header_map["Enabled"]).value = second_enabled

    workbook.save(template_path)
    return template_path, fields


def test_given_enabled_expected_event_when_actual_event_matches_then_case_is_ok(
    tmp_path: Path,
) -> None:
    schema_config = _schema_config()
    template_path, fields = _write_case_sheet(tmp_path, first_subject="Subject-A")
    testcases = read_template(template_path, [field.path for field in fields]).testcases

    result = match_and_validate(
        to_expected_events(testcases),
        to_actual_events(
            [_kafka_message(sender="sender@example.com", subject="Subject-A", score=1.55)]
        ),
        MatchingConfig(from_field="sender", subject_field="subject"),
        fields,
    )

    assert len(result.matches) == 1
    assert result.matches[0].expected_event.expected_event_id == "TC-1"
    assert result.matches[0].mismatches == ()
    assert result.conflicts == ()
    assert result.unmatched_expected_event_ids == ()
    assert schema_config.schema_type == "json_schema"


def test_given_sender_collision_when_subject_matches_one_case_then_single_case_is_selected(
    tmp_path: Path,
) -> None:
    template_path, fields = _write_case_sheet(
        tmp_path,
        first_subject="Subject-A",
        second_subject="Subject-B",
    )
    testcases = read_template(template_path, [field.path for field in fields]).testcases

    result = match_and_validate(
        to_expected_events(testcases),
        to_actual_events(
            [_kafka_message(sender="sender@example.com", subject="Subject-B", score=1.5)]
        ),
        MatchingConfig(from_field="sender", subject_field="subject"),
        fields,
    )

    assert len(result.matches) == 1
    assert result.matches[0].expected_event.expected_event_id == "TC-2"
    assert result.conflicts == ()
    assert result.unmatched_expected_event_ids == ("TC-1",)


def test_given_send_failure_when_writing_run_report_then_status_is_send_failed(
    tmp_path: Path,
) -> None:
    schema_config = _schema_config()
    template_path, fields = _write_case_sheet(tmp_path, first_subject="Subject-A")
    testcases = read_template(template_path, [field.path for field in fields]).testcases
    output_path = tmp_path / "results.xlsx"

    write_results_workbook(
        template_path=template_path,
        output_path=output_path,
        schema_config=schema_config,
        schema_fields=fields,
        testcases=testcases,
        match_result=MatchValidationResult(
            matches=(),
            conflicts=(),
            unmatched_actual_events=(),
            unmatched_expected_event_ids=("TC-1",),
        ),
        run_metadata=RunMetadata(
            run_start=datetime(2026, 2, 23, 18, 0, tzinfo=UTC),
            input_path=template_path,
            output_path=output_path,
            kafka_topic="topic-a",
            timeout_seconds=600,
            sent_ok=0,
        ),
        send_status_by_test_id={"TC-1": SendStatus.FAILED},
    )

    workbook = load_workbook(output_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    assert sheet.cell(row=3, column=sheet.max_column).value == "SEND_FAILED"

    run_info_sheet = workbook["RunInfo"]
    run_info = {
        run_info_sheet.cell(row=row, column=1).value: run_info_sheet.cell(row=row, column=2).value
        for row in range(1, 20)
    }
    assert run_info["passed"] == 0
    assert run_info["failed"] == 1


def test_given_tolerance_expression_when_actual_value_crosses_boundaries_then_only_in_range_is_ok(
    tmp_path: Path,
) -> None:
    template_path, fields = _write_case_sheet(
        tmp_path,
        first_subject="Subject-A",
        expected_score="3,14+-0,2",
    )
    testcases = read_template(template_path, [field.path for field in fields]).testcases

    in_range_result = match_and_validate(
        to_expected_events(testcases),
        to_actual_events(
            [_kafka_message(sender="sender@example.com", subject="Subject-A", score=3.30)]
        ),
        MatchingConfig(from_field="sender", subject_field="subject"),
        fields,
    )
    out_of_range_result = match_and_validate(
        to_expected_events(testcases),
        to_actual_events(
            [_kafka_message(sender="sender@example.com", subject="Subject-A", score=3.50)]
        ),
        MatchingConfig(from_field="sender", subject_field="subject"),
        fields,
    )

    assert in_range_result.matches[0].mismatches == ()
    assert len(out_of_range_result.matches[0].mismatches) == 1


def test_given_generate_config_when_building_test_configuration_then_placeholders_are_rendered(
    tmp_path: Path,
) -> None:
    runner = CliRunner()
    output_path = tmp_path / "config.yaml"

    result = runner.invoke(cli, ["generate-config", "--output", str(output_path)])

    assert result.exit_code == 0
    template = output_path.read_text(encoding="utf-8")
    assert "Test configuration template" in template
    assert "event schema" in template
    assert "matching.from_field" in template
    assert "<REQUIRED>" in template
    assert "<OPTIONAL>" in template


def test_given_missing_smtp_parallelism_when_loading_test_configuration_then_default_is_four(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        """
schema:
  json_schema:
    inline: |
      {
        "type": "object",
        "properties": {
          "sender": {"type": "string"},
          "subject": {"type": "string"}
        }
      }
matching:
  from_field: "sender"
  subject_field: "subject"
smtp:
  host: smtp.example.com
  port: 25
mail:
  to_address: "qa@example.com"
kafka:
  bootstrap_servers: "localhost:9092"
  topic: "result-topic"
""",
        encoding="utf-8",
    )

    configuration = load_configuration(config_path)

    assert configuration.smtp.parallelism == 4


def test_given_project_init_when_command_runs_then_local_venv_is_prepared(
    monkeypatch,
) -> None:
    runner = CliRunner()
    invoked_repo_roots: list[Path] = []

    def _fake_bootstrap(*, repo_root: Path) -> None:
        invoked_repo_roots.append(repo_root)

    monkeypatch.setattr("simple_e2e_tester.cli.bootstrap_project_environment", _fake_bootstrap)

    result = runner.invoke(cli, ["init"])

    assert result.exit_code == 0
    assert "local virtual environment ready" in result.output
    assert len(invoked_repo_roots) == 1


def test_given_synced_project_when_running_python_e2k_launcher_then_help_lists_commands() -> None:
    project_root = Path(__file__).resolve().parents[3]

    result = subprocess.run(
        [sys.executable, "e2k-tester", "--help"],
        cwd=project_root,
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0
    assert "init" in result.stdout
    assert "bootstrap" not in result.stdout
    assert "generate-config" in result.stdout
    assert "generate-template" in result.stdout
    assert "run" in result.stdout


def test_given_python3_when_running_launcher_then_help_lists_commands() -> None:
    project_root = Path(__file__).resolve().parents[3]
    python3_binary = shutil.which("python3")
    if python3_binary is None:
        pytest.skip("python3 not available in PATH")

    result = subprocess.run(
        [python3_binary, "e2k-tester", "--help"],
        cwd=project_root,
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0
    assert "init" in result.stdout
    assert "bootstrap" not in result.stdout
    assert "generate-config" in result.stdout
    assert "generate-template" in result.stdout
    assert "run" in result.stdout


def test_given_run_command_help_when_rendering_options_then_dry_run_is_supported() -> None:
    runner = CliRunner()

    result = runner.invoke(cli, ["run", "--help"])

    assert result.exit_code == 0
    assert "--dry-run" in result.output
    assert "not part of spec" not in result.output.lower()


def test_given_json_payload_when_consuming_with_json_schema_then_fields_match() -> None:
    schema_config = _schema_config()
    schema_fields = flatten_schema(load_schema_document(schema_config))
    start_time = datetime.now(UTC)
    payload = json.dumps(
        {"sender": "sender@example.com", "subject": "Subject-1", "score": 1.55}
    ).encode("utf-8")

    class _FakeRecord:
        def error(self) -> None:
            return None

        def key(self) -> None:
            return None

        def value(self) -> bytes:
            return payload

        def timestamp(self) -> tuple[int, int]:
            return (0, int((start_time.timestamp() + 1) * 1000))

    class _FakeConsumer:
        def __init__(self) -> None:
            self._served = False

        def subscribe(
            self,
            topics: list[str],
            on_assign: object | None = None,
            on_revoke: object | None = None,
            on_lost: object | None = None,
        ) -> None:
            return None

        def poll(self, timeout: float) -> _FakeRecord | None:
            if self._served:
                return None
            self._served = True
            return _FakeRecord()

        def close(self) -> None:
            return None

    reader = ActualEventReader(
        kafka_settings=KafkaSettings(
            bootstrap_servers=("localhost:9092",),
            topic="result-topic",
            group_id="group",
            security={},
            timeout_seconds=1,
            poll_interval_ms=10,
            auto_offset_reset="latest",
        ),
        schema_fields=schema_fields,
        schema_config=schema_config,
        consumer=_FakeConsumer(),
    )

    actual_events = list(reader.consume_from(start_time))

    assert len(actual_events) == 1
    assert actual_events[0].flattened["sender"] == "sender@example.com"
    assert actual_events[0].flattened["subject"] == "Subject-1"
    assert actual_events[0].flattened["score"] == 1.55


def test_given_all_enabled_expected_events_when_live_run_matches_then_kafka_reading_stops_early(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.json"
    config_path.write_text(
        json.dumps(
            {
                "schema": {"json_schema": {"inline": _schema_config().text}},
                "matching": {"from_field": "sender", "subject_field": "subject"},
                "smtp": {"host": "smtp.example.com", "port": 25},
                "mail": {"to_address": "qa@example.com"},
                "kafka": {
                    "bootstrap_servers": "localhost:9092",
                    "topic": "result-topic",
                    "timeout_seconds": 600,
                    "poll_interval_ms": 5,
                },
            }
        ),
        encoding="utf-8",
    )
    template_path, _ = _write_case_sheet(
        tmp_path,
        first_subject="Subject-A",
        second_subject="Subject-B",
    )
    output_dir = tmp_path / "results"
    trailing_events_consumed = 0

    class _FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [EmailSendResult.sent(testcase.test_id) for testcase in testcases]

    class _FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            nonlocal trailing_events_consumed
            assert isinstance(start_time, datetime)
            yield _kafka_message(sender="sender@example.com", subject="Subject-A", score=1.5)
            yield _kafka_message(sender="sender@example.com", subject="Subject-B", score=1.5)
            for index in range(3):
                trailing_events_consumed += 1
                yield _kafka_message(
                    sender="sender@example.com",
                    subject=f"Trailing-{index}",
                    score=1.5,
                )

    outcome = execute_email_kafka_validation_run(
        RunRequest(
            config_path=str(config_path),
            input_path=str(template_path),
            output_dir=str(output_dir),
            dry_run=False,
        ),
        email_sender_cls=_FakeEmailSender,
        kafka_service_cls=_FakeKafkaService,
    )

    assert outcome.sent_ok == 2
    assert trailing_events_consumed == 0


def test_given_no_enabled_test_case_when_live_run_executes_then_actual_event_consumption_is_skipped(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.json"
    config_path.write_text(
        json.dumps(
            {
                "schema": {"json_schema": {"inline": _schema_config().text}},
                "matching": {"from_field": "sender", "subject_field": "subject"},
                "smtp": {"host": "smtp.example.com", "port": 25},
                "mail": {"to_address": "qa@example.com"},
                "kafka": {
                    "bootstrap_servers": "localhost:9092",
                    "topic": "result-topic",
                    "timeout_seconds": 600,
                    "poll_interval_ms": 5,
                },
            }
        ),
        encoding="utf-8",
    )
    template_path, _ = _write_case_sheet(
        tmp_path,
        first_subject="Subject-A",
        first_enabled=False,
    )
    output_dir = tmp_path / "results"
    consume_calls = 0

    class _FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [EmailSendResult.skipped(testcase.test_id) for testcase in testcases]

    class _FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            nonlocal consume_calls
            consume_calls += 1
            return iter(())

    outcome = execute_email_kafka_validation_run(
        RunRequest(
            config_path=str(config_path),
            input_path=str(template_path),
            output_dir=str(output_dir),
            dry_run=False,
        ),
        email_sender_cls=_FakeEmailSender,
        kafka_service_cls=_FakeKafkaService,
    )

    assert outcome.sent_ok == 0
    assert consume_calls == 0
