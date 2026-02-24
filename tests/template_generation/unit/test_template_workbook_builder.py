"""Template generation service tests."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook
from simple_e2e_tester.configuration.runtime_settings import SchemaConfig
from simple_e2e_tester.schema_management import flatten_schema, load_schema_document
from simple_e2e_tester.template_generation.template_workbook_builder import (
    TEMPLATE_SHEET_NAME,
    generate_template_workbook,
)


def _build_schema_config() -> SchemaConfig:
    schema_text = json.dumps(
        {
            "type": "object",
            "properties": {
                "emailabsender": {"type": "string"},
                "emailbetreff": {"type": "string"},
                "ki_ergebnis": {
                    "type": "object",
                    "properties": {
                        "klasse": {"type": "array", "items": {"type": "string"}},
                        "fachdaten": {
                            "type": "object",
                            "properties": {
                                "grund": {
                                    "type": "object",
                                    "properties": {
                                        "value": {"type": "string"},
                                        "score": {"type": "number"},
                                    },
                                }
                            },
                        },
                    },
                },
            },
        }
    )
    return SchemaConfig(schema_type="json_schema", text=schema_text, source_path=None)


def test_template_contains_expected_columns_and_groups(tmp_path: Path) -> None:
    schema_config = _build_schema_config()
    schema_document = load_schema_document(schema_config)
    fields = flatten_schema(schema_document)
    output_path = tmp_path / "template.xlsx"

    generate_template_workbook(schema_config, fields, output_path)

    assert output_path.exists()
    workbook = load_workbook(output_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]

    expected_field_names = [field.path for field in fields]
    metadata_columns = ["ID", "Tags", "Enabled", "Notes"]
    input_columns = ["FROM", "SUBJECT", "BODY", "ATTACHMENT"]
    header_row = [
        sheet.cell(row=2, column=idx + 1).value
        for idx in range(len(metadata_columns + input_columns + expected_field_names))
    ]

    assert header_row == metadata_columns + input_columns + expected_field_names

    merged_ranges = {str(rng) for rng in sheet.merged_cells.ranges}
    assert "A1:D1" in merged_ranges
    assert "E1:H1" in merged_ranges
    last_col_letter = sheet.cell(row=2, column=len(header_row)).column_letter
    assert f"I1:{last_col_letter}1" in merged_ranges


def test_schema_sheet_contains_hash_and_text(tmp_path: Path) -> None:
    schema_config = _build_schema_config()
    schema_document = load_schema_document(schema_config)
    fields = flatten_schema(schema_document)
    output_path = tmp_path / "template.xlsx"

    generate_template_workbook(schema_config, fields, output_path)

    workbook = load_workbook(output_path, data_only=True)
    schema_sheet = workbook["Schema"]

    schema_hash = hashlib.sha256(schema_config.text.encode("utf-8")).hexdigest()
    assert schema_sheet["A1"].value == "schema_type"
    assert schema_sheet["B1"].value == schema_config.schema_type
    assert schema_sheet["A2"].value == "schema_hash"
    assert schema_sheet["B2"].value == schema_hash
    assert schema_sheet["A3"].value == "schema_text"
    assert schema_sheet["B3"].value == schema_config.text
