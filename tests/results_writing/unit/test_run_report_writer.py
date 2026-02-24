"""Results workbook writer tests."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from simple_e2e_tester.configuration.runtime_settings import SchemaConfig
from simple_e2e_tester.email_sending.delivery_outcomes import SendStatus
from simple_e2e_tester.kafka_consumption.actual_event_messages import ActualEventMessage
from simple_e2e_tester.matching_validation.event_boundary_mappers import (
    to_actual_events,
    to_expected_events,
)
from simple_e2e_tester.matching_validation.matching_outcomes import (
    FieldMismatch,
    MatchingConflict,
    MatchValidationResult,
    ValidatedMatch,
)
from simple_e2e_tester.results_writing.run_report_writer import RunMetadata, write_results_workbook
from simple_e2e_tester.schema_management import flatten_schema, load_schema_document
from simple_e2e_tester.template_generation import TEMPLATE_SHEET_NAME, generate_template_workbook
from simple_e2e_tester.template_ingestion.workbook_reader import read_template


def _schema_config() -> SchemaConfig:
    return SchemaConfig(
        schema_type="json_schema",
        text=json.dumps(
            {
                "type": "object",
                "properties": {
                    "sender": {"type": "string"},
                    "subject": {"type": "string"},
                    "score": {"type": "number"},
                },
            }
        ),
        source_path=None,
    )


def _prepare_template(tmp_path: Path) -> tuple[Path, list]:
    schema_config = _schema_config()
    fields = flatten_schema(load_schema_document(schema_config))
    template_path = tmp_path / "input.xlsx"
    generate_template_workbook(schema_config, fields, template_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }

    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Subject-1"
    sheet.cell(row=3, column=header_map["score"]).value = "1,50+-0,1"

    sheet.cell(row=4, column=header_map["ID"]).value = "TC-2"
    sheet.cell(row=4, column=header_map["FROM"]).value = "sender2@example.com"
    sheet.cell(row=4, column=header_map["SUBJECT"]).value = "Subject-2"
    sheet.cell(row=4, column=header_map["score"]).value = "2,00"
    workbook.save(template_path)
    return template_path, fields


def _message(sender: str, subject: str, score: float) -> ActualEventMessage:
    flattened = {"sender": sender, "subject": subject, "score": score}
    return ActualEventMessage(
        key=None,
        value=flattened,
        timestamp=datetime.now(UTC),
        flattened=flattened,
    )


def test_writes_output_with_actual_match_and_duplicated_rows(tmp_path: Path) -> None:
    template_path, fields = _prepare_template(tmp_path)
    schema_config = _schema_config()
    testcases = read_template(template_path, [field.path for field in fields]).testcases
    tc1 = next(case for case in testcases if case.test_id == "TC-1")
    expected_tc1 = to_expected_events([tc1])[0]

    first_match = ValidatedMatch(
        expected_event=expected_tc1,
        actual_event=to_actual_events([_message("sender@example.com", "Subject-1", 1.55)])[0],
        mismatches=(),
    )
    second_match = ValidatedMatch(
        expected_event=expected_tc1,
        actual_event=to_actual_events([_message("sender@example.com", "Subject-1", 1.80)])[0],
        mismatches=(FieldMismatch(field="score", expected="1,50+-0,1", actual="1.8"),),
    )
    result = MatchValidationResult(
        matches=(first_match, second_match),
        conflicts=(),
        unmatched_actual_events=(),
        unmatched_expected_event_ids=("TC-2",),
    )

    output_path = tmp_path / "results.xlsx"
    write_results_workbook(
        template_path=template_path,
        output_path=output_path,
        schema_config=schema_config,
        schema_fields=fields,
        testcases=testcases,
        match_result=result,
        run_metadata=RunMetadata(
            run_start=datetime(2026, 2, 23, 12, 0, tzinfo=UTC),
            input_path=template_path,
            output_path=output_path,
            kafka_topic="topic-a",
            timeout_seconds=600,
            sent_ok=1,
        ),
    )

    workbook = load_workbook(output_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    assert sheet.cell(row=1, column=1).value == "Metadata"
    assert sheet.cell(row=1, column=5).value == "Input"
    assert sheet.cell(row=1, column=9).value == "Expected"
    assert sheet.cell(row=1, column=12).value == "Actual"
    assert sheet.cell(row=1, column=15).value == "Match"
    assert sheet.cell(row=2, column=15).value == "Match"

    assert sheet.cell(row=3, column=1).value == "TC-1"
    assert sheet.cell(row=4, column=1).value == "TC-1"
    assert sheet.cell(row=5, column=1).value == "TC-2"

    assert sheet.cell(row=3, column=14).value == 1.55
    assert sheet.cell(row=3, column=15).value == "OK"
    assert sheet.cell(row=4, column=14).value == 1.8
    assert "expected: 1,50+-0,1" in str(sheet.cell(row=4, column=15).value)
    assert "actual: 1.8" in str(sheet.cell(row=4, column=15).value)
    assert sheet.cell(row=5, column=15).value == "NOT_FOUND"

    assert "Schema" in workbook.sheetnames
    assert "RunInfo" in workbook.sheetnames
    run_info_sheet = workbook["RunInfo"]
    run_info = {
        run_info_sheet.cell(row=row, column=1).value: run_info_sheet.cell(row=row, column=2).value
        for row in range(1, 20)
    }
    assert run_info["kafka_topic"] == "topic-a"
    assert run_info["matched"] == 2
    assert run_info["passed"] == 1
    assert run_info["failed"] == 1
    assert run_info["not_found"] == 1


def test_marks_conflict_and_send_failure_rows(tmp_path: Path) -> None:
    template_path, fields = _prepare_template(tmp_path)
    schema_config = _schema_config()
    testcases = read_template(template_path, [field.path for field in fields]).testcases
    result = MatchValidationResult(
        matches=(),
        conflicts=(
            MatchingConflict(
                actual_event=to_actual_events([_message("x@example.com", "x", 0.0)])[0],
                candidate_expected_event_ids=("TC-2",),
            ),
        ),
        unmatched_actual_events=(),
        unmatched_expected_event_ids=("TC-1", "TC-2"),
    )

    output_path = tmp_path / "results.xlsx"
    write_results_workbook(
        template_path=template_path,
        output_path=output_path,
        schema_config=schema_config,
        schema_fields=fields,
        testcases=testcases,
        match_result=result,
        run_metadata=RunMetadata(
            run_start=datetime(2026, 2, 23, 12, 0, tzinfo=UTC),
            input_path=template_path,
            output_path=output_path,
            kafka_topic="topic-a",
            timeout_seconds=600,
            sent_ok=0,
        ),
        send_status_by_test_id={"TC-1": SendStatus.FAILED},
    )

    sheet = load_workbook(output_path)[TEMPLATE_SHEET_NAME]
    assert sheet.cell(row=3, column=15).value == "SEND_FAILED"
    assert sheet.cell(row=4, column=15).value == "CONFLICT"


def test_not_found_excludes_skipped_rows(tmp_path: Path) -> None:
    template_path, fields = _prepare_template(tmp_path)
    schema_config = _schema_config()
    testcases = read_template(template_path, [field.path for field in fields]).testcases
    result = MatchValidationResult(
        matches=(),
        conflicts=(),
        unmatched_actual_events=(),
        unmatched_expected_event_ids=("TC-1",),
    )

    output_path = tmp_path / "results.xlsx"
    write_results_workbook(
        template_path=template_path,
        output_path=output_path,
        schema_config=schema_config,
        schema_fields=fields,
        testcases=testcases,
        match_result=result,
        run_metadata=RunMetadata(
            run_start=datetime(2026, 2, 23, 12, 0, tzinfo=UTC),
            input_path=template_path,
            output_path=output_path,
            kafka_topic="topic-a",
            timeout_seconds=600,
            sent_ok=0,
        ),
        send_status_by_test_id={"TC-1": SendStatus.SKIPPED},
    )

    workbook = load_workbook(output_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    assert sheet.cell(row=3, column=15).value == "SKIPPED"

    run_info_sheet = workbook["RunInfo"]
    run_info = {
        run_info_sheet.cell(row=row, column=1).value: run_info_sheet.cell(row=row, column=2).value
        for row in range(1, 20)
    }
    assert run_info["not_found"] == 0
