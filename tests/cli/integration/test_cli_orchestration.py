"""CLI orchestration integration tests."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

from click.testing import CliRunner
from openpyxl import load_workbook
from simple_e2e_tester.cli import cli
from simple_e2e_tester.email_sending.delivery_outcomes import EmailSendResult
from simple_e2e_tester.kafka_consumption.actual_event_messages import ActualEventMessage
from simple_e2e_tester.template_generation import TEMPLATE_SHEET_NAME


def _write_config(tmp_path: Path, schema_type: str = "json_schema") -> Path:
    json_schema = json.dumps(
        {
            "type": "object",
            "properties": {
                "sender": {"type": "string"},
                "subject": {"type": "string"},
                "score": {"type": "number"},
            },
        }
    )
    avsc_schema = json.dumps(
        {
            "type": "record",
            "name": "Root",
            "fields": [
                {"name": "sender", "type": "string"},
                {"name": "subject", "type": "string"},
                {"name": "score", "type": "double"},
            ],
        }
    )
    if schema_type == "json_schema":
        schema_config = {"json_schema": {"inline": json_schema}}
    elif schema_type == "avsc":
        schema_config = {"avsc": {"inline": avsc_schema}}
    else:
        raise ValueError(f"Unsupported schema_type for test helper: {schema_type}")

    config = {
        "schema": schema_config,
        "matching": {"from_field": "sender", "subject_field": "subject"},
        "smtp": {"host": "smtp.example.com", "port": 25},
        "mail": {"to_address": "qa@example.com"},
        "kafka": {"bootstrap_servers": "localhost:9092", "topic": "result-topic"},
    }
    path = tmp_path / "config.json"
    path.write_text(json.dumps(config), encoding="utf-8")
    return path


def test_generate_template_command_writes_workbook(tmp_path: Path) -> None:
    runner = CliRunner()
    config_path = _write_config(tmp_path)
    output_path = tmp_path / "generated-template.xlsx"

    result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(output_path),
        ],
    )

    assert result.exit_code == 0
    assert output_path.exists()
    workbook = load_workbook(output_path)
    assert TEMPLATE_SHEET_NAME in workbook.sheetnames
    assert "Schema" in workbook.sheetnames


def test_generate_template_command_returns_error_for_invalid_config(tmp_path: Path) -> None:
    runner = CliRunner()
    config_path = tmp_path / "invalid-config.json"
    config_path.write_text(json.dumps({"schema": {}, "smtp": {}}), encoding="utf-8")
    output_path = tmp_path / "generated-template.xlsx"

    result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(output_path),
        ],
    )

    assert result.exit_code != 0
    assert "Exactly one event schema type" in str(result.exception)
    assert not output_path.exists()


def test_generate_config_command_writes_placeholder_file_with_default_name(tmp_path: Path) -> None:
    runner = CliRunner()
    with runner.isolated_filesystem(temp_dir=str(tmp_path)):
        result = runner.invoke(cli, ["generate-config"])
        output_path = Path("config.yaml").resolve()

        assert result.exit_code == 0
        assert output_path.exists()
        content = output_path.read_text(encoding="utf-8")
        assert "schema:" in content
        assert "matching:" in content
        assert "smtp:" in content
        assert "mail:" in content
        assert "kafka:" in content
        assert "<REQUIRED>" in content
        assert "<OPTIONAL>" in content
        assert "# Choose exactly one event schema type" in content
        assert str(output_path) in result.output


def test_generate_config_command_fails_when_output_file_already_exists(tmp_path: Path) -> None:
    runner = CliRunner()
    output_path = tmp_path / "config.yaml"
    output_path.write_text("already-there", encoding="utf-8")

    result = runner.invoke(
        cli,
        [
            "generate-config",
            "--output",
            str(output_path),
        ],
    )

    assert result.exit_code != 0
    assert "already exists" in str(result.exception).lower()
    assert output_path.read_text(encoding="utf-8") == "already-there"


def test_run_command_dry_run_writes_results_workbook(tmp_path: Path) -> None:
    runner = CliRunner()
    config_path = _write_config(tmp_path)
    template_path = tmp_path / "generated-template.xlsx"
    output_dir = tmp_path / "results"

    generate_result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(template_path),
        ],
    )
    assert generate_result.exit_code == 0

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Subject-1"
    sheet.cell(row=3, column=header_map["score"]).value = "1,50"
    workbook.save(template_path)

    run_result = runner.invoke(
        cli,
        [
            "run",
            "--config",
            str(config_path),
            "--input",
            str(template_path),
            "--output-dir",
            str(output_dir),
            "--dry-run",
        ],
    )

    assert run_result.exit_code == 0
    result_files = list(output_dir.glob("*-results-*.xlsx"))
    assert len(result_files) == 1
    result_workbook = load_workbook(result_files[0])
    assert TEMPLATE_SHEET_NAME in result_workbook.sheetnames
    assert "RunInfo" in result_workbook.sheetnames

    result_sheet = result_workbook[TEMPLATE_SHEET_NAME]
    assert result_sheet.cell(row=2, column=result_sheet.max_column).value == "Match"
    assert result_sheet.cell(row=3, column=result_sheet.max_column).value == "SKIPPED"
    run_info_sheet = result_workbook["RunInfo"]
    run_info = {
        run_info_sheet.cell(row=row, column=1).value: run_info_sheet.cell(row=row, column=2).value
        for row in range(1, 20)
    }
    assert run_info["matched"] == 0
    assert run_info["failed"] == 0
    assert run_info["not_found"] == 0


def test_run_command_with_mocked_dependencies_writes_validated_results(
    tmp_path: Path, monkeypatch
) -> None:
    runner = CliRunner()
    config_path = _write_config(tmp_path, schema_type="avsc")
    template_path = tmp_path / "generated-template.xlsx"
    output_dir = tmp_path / "results"

    generate_result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(template_path),
        ],
    )
    assert generate_result.exit_code == 0

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Subject-1"
    sheet.cell(row=3, column=header_map["score"]).value = "1,50+-0,1"
    workbook.save(template_path)

    class FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [EmailSendResult.sent(testcase.test_id) for testcase in testcases]

    class FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            yield ActualEventMessage(
                key=None,
                value={},
                timestamp=datetime.now(UTC),
                flattened={
                    "sender": "sender@example.com",
                    "subject": "Subject-1",
                    "score": 1.58,
                },
            )

    monkeypatch.setattr("simple_e2e_tester.cli.ExpectedEventDispatcher", FakeEmailSender)
    monkeypatch.setattr("simple_e2e_tester.cli.ActualEventReader", FakeKafkaService)

    run_result = runner.invoke(
        cli,
        [
            "run",
            "--config",
            str(config_path),
            "--input",
            str(template_path),
            "--output-dir",
            str(output_dir),
        ],
    )
    assert run_result.exit_code == 0

    result_files = list(output_dir.glob("*-results-*.xlsx"))
    assert len(result_files) == 1
    result_workbook = load_workbook(result_files[0])
    result_sheet = result_workbook[TEMPLATE_SHEET_NAME]
    assert result_sheet.cell(row=3, column=result_sheet.max_column).value == "OK"

    run_info_sheet = result_workbook["RunInfo"]
    run_info = {
        run_info_sheet.cell(row=row, column=1).value: run_info_sheet.cell(row=row, column=2).value
        for row in range(1, 20)
    }
    assert run_info["sent_ok"] == 1
    assert run_info["matched"] == 1
    assert run_info["passed"] == 1


def test_run_command_marks_send_failures_in_output(tmp_path: Path, monkeypatch) -> None:
    runner = CliRunner()
    config_path = _write_config(tmp_path, schema_type="avsc")
    template_path = tmp_path / "generated-template.xlsx"
    output_dir = tmp_path / "results"

    generate_result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(template_path),
        ],
    )
    assert generate_result.exit_code == 0

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Subject-1"
    workbook.save(template_path)

    class FailingEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [
                EmailSendResult.failed(testcase.test_id, RuntimeError("SMTP unavailable"))
                for testcase in testcases
            ]

    class EmptyKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            return iter(())

    monkeypatch.setattr("simple_e2e_tester.cli.ExpectedEventDispatcher", FailingEmailSender)
    monkeypatch.setattr("simple_e2e_tester.cli.ActualEventReader", EmptyKafkaService)

    run_result = runner.invoke(
        cli,
        [
            "run",
            "--config",
            str(config_path),
            "--input",
            str(template_path),
            "--output-dir",
            str(output_dir),
        ],
    )
    assert run_result.exit_code == 0

    result_files = list(output_dir.glob("*-results-*.xlsx"))
    assert len(result_files) == 1
    result_workbook = load_workbook(result_files[0])
    result_sheet = result_workbook[TEMPLATE_SHEET_NAME]
    assert result_sheet.cell(row=3, column=result_sheet.max_column).value == "SEND_FAILED"

    run_info_sheet = result_workbook["RunInfo"]
    run_info = {
        run_info_sheet.cell(row=row, column=1).value: run_info_sheet.cell(row=row, column=2).value
        for row in range(1, 20)
    }
    assert run_info["failed"] == 1


def test_run_command_supports_json_schema_with_mocked_dependencies(
    tmp_path: Path, monkeypatch
) -> None:
    runner = CliRunner()
    config_path = _write_config(tmp_path, schema_type="json_schema")
    template_path = tmp_path / "generated-template.xlsx"
    output_dir = tmp_path / "results"

    generate_result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(template_path),
        ],
    )
    assert generate_result.exit_code == 0

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Subject-1"
    workbook.save(template_path)

    class FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [EmailSendResult.sent(testcase.test_id) for testcase in testcases]

    class FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            yield ActualEventMessage(
                key=None,
                value={},
                timestamp=datetime.now(UTC),
                flattened={
                    "sender": "sender@example.com",
                    "subject": "Subject-1",
                    "score": 1.58,
                },
            )

    monkeypatch.setattr("simple_e2e_tester.cli.ExpectedEventDispatcher", FakeEmailSender)
    monkeypatch.setattr("simple_e2e_tester.cli.ActualEventReader", FakeKafkaService)

    run_result = runner.invoke(
        cli,
        [
            "run",
            "--config",
            str(config_path),
            "--input",
            str(template_path),
            "--output-dir",
            str(output_dir),
        ],
    )

    assert run_result.exit_code == 0
    result_files = list(output_dir.glob("*-results-*.xlsx"))
    assert len(result_files) == 1
    result_workbook = load_workbook(result_files[0])
    result_sheet = result_workbook[TEMPLATE_SHEET_NAME]
    assert result_sheet.cell(row=3, column=result_sheet.max_column).value == "OK"


def test_run_command_validates_attachment_paths_before_sender_initialization(
    tmp_path: Path, monkeypatch
) -> None:
    runner = CliRunner()
    config_path = _write_config(tmp_path, schema_type="avsc")
    template_path = tmp_path / "generated-template.xlsx"

    generate_result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(template_path),
        ],
    )
    assert generate_result.exit_code == 0

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Subject-1"
    sheet.cell(row=3, column=header_map["ATTACHMENT"]).value = "./missing.txt"
    workbook.save(template_path)

    state = {"sender_initialized": False}

    class GuardEmailSender:
        def __init__(self, **kwargs) -> None:
            state["sender_initialized"] = True
            raise AssertionError("email sender must not be initialized")

    monkeypatch.setattr("simple_e2e_tester.cli.ExpectedEventDispatcher", GuardEmailSender)

    run_result = runner.invoke(
        cli,
        [
            "run",
            "--config",
            str(config_path),
            "--input",
            str(template_path),
        ],
    )

    assert run_result.exit_code != 0
    assert state["sender_initialized"] is False
    assert "attachment file not found" in str(run_result.exception).lower()
