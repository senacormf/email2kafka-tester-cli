"""Tests for run execution use-case service."""

from __future__ import annotations

import json
import threading
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from simple_e2e_tester.email_sending.delivery_outcomes import EmailSendResult
from simple_e2e_tester.kafka_consumption.actual_event_messages import ActualEventMessage
from simple_e2e_tester.run_execution.run_contracts import RunRequest
from simple_e2e_tester.run_execution.validation_run_use_case import (
    execute_email_kafka_validation_run,
)
from simple_e2e_tester.template_generation import TEMPLATE_SHEET_NAME


def _write_config(tmp_path: Path, schema_type: str = "json_schema") -> Path:
    json_schema = json.dumps(
        {
            "type": "object",
            "properties": {
                "sender": {"type": "string"},
                "subject": {"type": "string"},
                "score": {"type": "number"},
            },
        }
    )
    avsc_schema = json.dumps(
        {
            "type": "record",
            "name": "Root",
            "fields": [
                {"name": "sender", "type": "string"},
                {"name": "subject", "type": "string"},
                {"name": "score", "type": "double"},
            ],
        }
    )
    schema_config = {"json_schema": {"inline": json_schema}}
    if schema_type == "avsc":
        schema_config = {"avsc": {"inline": avsc_schema}}

    config = {
        "schema": schema_config,
        "matching": {"from_field": "sender", "subject_field": "subject"},
        "smtp": {"host": "smtp.example.com", "port": 25},
        "mail": {"to_address": "qa@example.com"},
        "kafka": {"bootstrap_servers": "localhost:9092", "topic": "result-topic"},
    }
    path = tmp_path / "config.json"
    path.write_text(json.dumps(config), encoding="utf-8")
    return path


def _write_template(
    tmp_path: Path,
    config_path: Path,
    *,
    second_subject: str | None = None,
) -> Path:
    from click.testing import CliRunner
    from simple_e2e_tester.cli import cli

    template_path = tmp_path / "generated-template.xlsx"
    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(template_path),
        ],
    )
    assert result.exit_code == 0

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Subject-1"
    sheet.cell(row=3, column=header_map["score"]).value = "1,50+-0,1"
    if second_subject is not None:
        sheet.cell(row=4, column=header_map["ID"]).value = "TC-2"
        sheet.cell(row=4, column=header_map["FROM"]).value = "sender@example.com"
        sheet.cell(row=4, column=header_map["SUBJECT"]).value = second_subject
        sheet.cell(row=4, column=header_map["score"]).value = "1,50+-0,1"
    workbook.save(template_path)
    return template_path


def test_execute_run_use_case_dry_run_writes_results_and_returns_outcome(tmp_path: Path) -> None:
    config_path = _write_config(tmp_path, schema_type="json_schema")
    template_path = _write_template(tmp_path, config_path)
    output_dir = tmp_path / "results"

    outcome = execute_email_kafka_validation_run(
        RunRequest(
            config_path=str(config_path),
            input_path=str(template_path),
            output_dir=str(output_dir),
            dry_run=True,
        )
    )

    assert outcome.dry_run is True
    assert outcome.sent_ok == 0
    assert outcome.output_path.exists()

    result_sheet = load_workbook(outcome.output_path)[TEMPLATE_SHEET_NAME]
    assert result_sheet.cell(row=3, column=result_sheet.max_column).value == "SKIPPED"


def test_given_no_enabled_test_case_when_live_run_executes_then_kafka_reader_is_not_called(
    tmp_path: Path,
) -> None:
    config_path = _write_config(tmp_path, schema_type="json_schema")
    template_path = _write_template(tmp_path, config_path)
    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["Enabled"]).value = False
    workbook.save(template_path)
    output_dir = tmp_path / "results"
    consume_calls = 0

    class FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [EmailSendResult.skipped(testcase.test_id) for testcase in testcases]

    class FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            nonlocal consume_calls
            consume_calls += 1
            return iter(())

    outcome = execute_email_kafka_validation_run(
        RunRequest(
            config_path=str(config_path),
            input_path=str(template_path),
            output_dir=str(output_dir),
            dry_run=False,
        ),
        email_sender_cls=FakeEmailSender,
        kafka_service_cls=FakeKafkaService,
    )

    assert outcome.sent_ok == 0
    assert consume_calls == 0


def test_execute_run_use_case_live_mode_uses_injected_sender_and_kafka(
    tmp_path: Path, monkeypatch
) -> None:
    config_path = _write_config(tmp_path, schema_type="avsc")
    template_path = _write_template(tmp_path, config_path)
    output_dir = tmp_path / "results"

    class FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [EmailSendResult.sent(testcase.test_id) for testcase in testcases]

    class FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            assert isinstance(start_time, datetime)
            flattened = {
                "sender": "sender@example.com",
                "subject": "Subject-1",
                "score": 1.58,
            }
            yield ActualEventMessage(
                key=None,
                value=flattened,
                timestamp=datetime.now(UTC),
                flattened=flattened,
            )

    monkeypatch.setattr(
        "simple_e2e_tester.run_execution.validation_run_use_case.ExpectedEventDispatcher",
        FakeEmailSender,
    )
    monkeypatch.setattr(
        "simple_e2e_tester.run_execution.validation_run_use_case.ActualEventReader",
        FakeKafkaService,
    )

    outcome = execute_email_kafka_validation_run(
        RunRequest(
            config_path=str(config_path),
            input_path=str(template_path),
            output_dir=str(output_dir),
            dry_run=False,
        )
    )

    assert outcome.dry_run is False
    assert outcome.sent_ok == 1
    result_sheet = load_workbook(outcome.output_path)[TEMPLATE_SHEET_NAME]
    assert result_sheet.cell(row=3, column=result_sheet.max_column).value == "OK"


def test_execute_run_use_case_live_mode_accepts_json_schema_with_injected_services(
    tmp_path: Path,
) -> None:
    config_path = _write_config(tmp_path, schema_type="json_schema")
    template_path = _write_template(tmp_path, config_path)
    output_dir = tmp_path / "results"

    class FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [EmailSendResult.sent(testcase.test_id) for testcase in testcases]

    class FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            assert isinstance(start_time, datetime)
            flattened = {
                "sender": "sender@example.com",
                "subject": "Subject-1",
                "score": 1.58,
            }
            yield ActualEventMessage(
                key=None,
                value=flattened,
                timestamp=datetime.now(UTC),
                flattened=flattened,
            )

    outcome = execute_email_kafka_validation_run(
        RunRequest(
            config_path=str(config_path),
            input_path=str(template_path),
            output_dir=str(output_dir),
            dry_run=False,
        ),
        email_sender_cls=FakeEmailSender,
        kafka_service_cls=FakeKafkaService,
    )

    assert outcome.dry_run is False
    assert outcome.sent_ok == 1
    result_sheet = load_workbook(outcome.output_path)[TEMPLATE_SHEET_NAME]
    assert result_sheet.cell(row=3, column=result_sheet.max_column).value == "OK"


def test_execute_run_use_case_starts_actual_event_reading_while_sending(
    tmp_path: Path,
) -> None:
    config_path = _write_config(tmp_path, schema_type="avsc")
    template_path = _write_template(tmp_path, config_path)
    output_dir = tmp_path / "results"

    reader_started = threading.Event()
    sender_finished = threading.Event()
    reader_started_during_send = False

    class FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            nonlocal reader_started_during_send
            reader_started_during_send = reader_started.wait(timeout=0.3)
            sender_finished.set()
            return [EmailSendResult.sent(testcase.test_id) for testcase in testcases]

    class FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            assert isinstance(start_time, datetime)
            reader_started.set()
            sender_finished.wait(timeout=1.0)
            flattened = {
                "sender": "sender@example.com",
                "subject": "Subject-1",
                "score": 1.58,
            }
            yield ActualEventMessage(
                key=None,
                value=flattened,
                timestamp=datetime.now(UTC),
                flattened=flattened,
            )

    outcome = execute_email_kafka_validation_run(
        RunRequest(
            config_path=str(config_path),
            input_path=str(template_path),
            output_dir=str(output_dir),
            dry_run=False,
        ),
        email_sender_cls=FakeEmailSender,
        kafka_service_cls=FakeKafkaService,
    )

    assert outcome.sent_ok == 1
    assert reader_started_during_send is True


def test_given_all_enabled_expected_events_when_live_run_matches_then_kafka_reading_stops_early(
    tmp_path: Path,
) -> None:
    config_path = _write_config(tmp_path, schema_type="json_schema")
    template_path = _write_template(tmp_path, config_path, second_subject="Subject-2")
    output_dir = tmp_path / "results"
    trailing_events_consumed = 0

    class FakeEmailSender:
        def __init__(self, **kwargs) -> None:
            pass

        def send_all(self, testcases):
            return [EmailSendResult.sent(testcase.test_id) for testcase in testcases]

    class FakeKafkaService:
        def __init__(self, **kwargs) -> None:
            pass

        def consume_from(self, start_time):
            nonlocal trailing_events_consumed
            assert isinstance(start_time, datetime)
            flattened_first = {
                "sender": "sender@example.com",
                "subject": "Subject-1",
                "score": 1.58,
            }
            yield ActualEventMessage(
                key=None,
                value=flattened_first,
                timestamp=datetime.now(UTC),
                flattened=flattened_first,
            )
            flattened_second = {
                "sender": "sender@example.com",
                "subject": "Subject-2",
                "score": 1.52,
            }
            yield ActualEventMessage(
                key=None,
                value=flattened_second,
                timestamp=datetime.now(UTC),
                flattened=flattened_second,
            )
            for index in range(4):
                trailing_events_consumed += 1
                trailing = {
                    "sender": "sender@example.com",
                    "subject": f"Trailing-{index}",
                    "score": 2.0,
                }
                yield ActualEventMessage(
                    key=None,
                    value=trailing,
                    timestamp=datetime.now(UTC),
                    flattened=trailing,
                )

    outcome = execute_email_kafka_validation_run(
        RunRequest(
            config_path=str(config_path),
            input_path=str(template_path),
            output_dir=str(output_dir),
            dry_run=False,
        ),
        email_sender_cls=FakeEmailSender,
        kafka_service_cls=FakeKafkaService,
    )

    assert outcome.sent_ok == 2
    assert trailing_events_consumed == 0
