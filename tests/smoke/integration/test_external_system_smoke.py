"""Smoke tests with mocked SMTP/Kafka surrounding systems."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from email.message import EmailMessage
from pathlib import Path
from typing import ClassVar, cast

from click.testing import CliRunner
from openpyxl import load_workbook
from simple_e2e_tester.cli import cli
from simple_e2e_tester.configuration.runtime_settings import KafkaSettings
from simple_e2e_tester.kafka_consumption.observed_event_messages import ObservedEventMessage
from simple_e2e_tester.template_generation import TEMPLATE_SHEET_NAME


def _write_config(tmp_path: Path) -> Path:
    config = {
        "schema": {
            "avsc": {
                "inline": json.dumps(
                    {
                        "type": "record",
                        "name": "Root",
                        "fields": [
                            {"name": "sender", "type": "string"},
                            {"name": "subject", "type": "string"},
                            {"name": "score", "type": "double"},
                        ],
                    }
                )
            }
        },
        "matching": {"from_field": "sender", "subject_field": "subject"},
        "smtp": {
            "host": "smtp.example.com",
            "port": 587,
            "username": "smtp-user",
            "password": "smtp-pass",
            "use_starttls": True,
            "use_ssl": False,
        },
        "mail": {
            "to_address": "qa@example.com",
            "cc": ["copy@example.com"],
            "bcc": ["hidden@example.com"],
        },
        "kafka": {
            "bootstrap_servers": "broker-1:9092",
            "topic": "result-topic",
            "security": {
                "security.protocol": "SASL_SSL",
                "sasl.mechanism": "PLAIN",
                "sasl.username": "kafka-user",
                "sasl.password": "kafka-pass",
            },
        },
    }
    path = tmp_path / "config.json"
    path.write_text(json.dumps(config), encoding="utf-8")
    return path


class _FakeSmtpSession:
    instances: ClassVar[list[_FakeSmtpSession]] = []

    def __init__(self, host: str, port: int, timeout: int) -> None:
        self.host = host
        self.port = port
        self.timeout = timeout
        self.actions: list[object] = []
        self.sent_message: EmailMessage | None = None
        self.to_addrs: list[str] | None = None
        _FakeSmtpSession.instances.append(self)

    def ehlo(self) -> None:
        self.actions.append("ehlo")

    def starttls(self) -> None:
        self.actions.append("starttls")

    def login(self, username: str, password: str) -> None:
        self.actions.append(("login", username, password))

    def send_message(self, message: EmailMessage, to_addrs: list[str]) -> None:
        self.actions.append("send_message")
        self.sent_message = message
        self.to_addrs = list(to_addrs)

    def quit(self) -> None:
        self.actions.append("quit")


class _FakeKafkaService:
    last_kwargs: ClassVar[dict[str, object]] = {}
    last_start_time: ClassVar[datetime | None] = None

    def __init__(self, **kwargs) -> None:
        _FakeKafkaService.last_kwargs = kwargs

    def consume_from(self, start_time: datetime):
        _FakeKafkaService.last_start_time = start_time
        flattened = {
            "sender": "sender@example.com",
            "subject": "Subject-1",
            "score": 1.58,
        }
        return iter(
            [
                ObservedEventMessage(
                    key=None,
                    value=flattened,
                    timestamp=datetime.now(UTC),
                    flattened=flattened,
                )
            ]
        )


def test_run_smoke_mocks_smtp_and_kafka_and_validates_requests_and_results(
    tmp_path: Path, monkeypatch
) -> None:
    runner = CliRunner()
    config_path = _write_config(tmp_path)
    template_path = tmp_path / "template.xlsx"
    output_dir = tmp_path / "results"

    generate_result = runner.invoke(
        cli,
        [
            "generate-template",
            "--config",
            str(config_path),
            "--output",
            str(template_path),
        ],
    )
    assert generate_result.exit_code == 0

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    header_map = {
        sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=header_map["ID"]).value = "TC-1"
    sheet.cell(row=3, column=header_map["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=header_map["SUBJECT"]).value = "Subject-1"
    sheet.cell(row=3, column=header_map["score"]).value = "1,50+-0,1"
    workbook.save(template_path)

    _FakeSmtpSession.instances.clear()
    monkeypatch.setattr(
        "simple_e2e_tester.email_sending.email_dispatch.smtplib.SMTP",
        _FakeSmtpSession,
    )
    monkeypatch.setattr("simple_e2e_tester.cli.ObservedEventReader", _FakeKafkaService)

    run_result = runner.invoke(
        cli,
        [
            "run",
            "--config",
            str(config_path),
            "--input",
            str(template_path),
            "--output-dir",
            str(output_dir),
        ],
    )
    assert run_result.exit_code == 0

    assert len(_FakeSmtpSession.instances) == 1
    smtp_session = _FakeSmtpSession.instances[0]
    assert smtp_session.host == "smtp.example.com"
    assert smtp_session.port == 587
    assert smtp_session.timeout == 30
    assert smtp_session.actions == [
        "ehlo",
        "starttls",
        "ehlo",
        ("login", "smtp-user", "smtp-pass"),
        "send_message",
        "quit",
    ]
    assert smtp_session.to_addrs == [
        "qa@example.com",
        "copy@example.com",
        "hidden@example.com",
    ]
    assert smtp_session.sent_message is not None
    assert smtp_session.sent_message["From"] == "sender@example.com"
    assert smtp_session.sent_message["To"] == "qa@example.com"
    assert smtp_session.sent_message["Cc"] == "copy@example.com"
    assert smtp_session.sent_message["Bcc"] == "hidden@example.com"
    assert smtp_session.sent_message["Subject"] == "Subject-1"
    assert smtp_session.sent_message["X-Test-Id"] == "TC-1"

    kafka_settings = cast(KafkaSettings, _FakeKafkaService.last_kwargs["kafka_settings"])
    assert kafka_settings.topic == "result-topic"
    assert kafka_settings.security["security.protocol"] == "SASL_SSL"
    assert kafka_settings.security["sasl.username"] == "kafka-user"
    assert _FakeKafkaService.last_start_time is not None

    result_files = list(output_dir.glob("*-results-*.xlsx"))
    assert len(result_files) == 1
    result_workbook = load_workbook(result_files[0])
    result_sheet = result_workbook[TEMPLATE_SHEET_NAME]
    assert result_sheet.cell(row=3, column=result_sheet.max_column).value == "OK"
    assert result_sheet.cell(row=3, column=result_sheet.max_column - 1).value == 1.58

    run_info_sheet = result_workbook["RunInfo"]
    run_info = {
        run_info_sheet.cell(row=row, column=1).value: run_info_sheet.cell(row=row, column=2).value
        for row in range(1, 20)
    }
    assert run_info["sent_ok"] == 1
    assert run_info["matched"] == 1
    assert run_info["passed"] == 1
