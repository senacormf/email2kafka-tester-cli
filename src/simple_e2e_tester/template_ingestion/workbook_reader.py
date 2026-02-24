"""Template ingestion and validation service."""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from simple_e2e_tester.template_generation import (
    INPUT_COLUMNS,
    METADATA_COLUMNS,
    TEMPLATE_SHEET_NAME,
)

from .testcase_models import TemplateReadResult, TemplateTestCase

EMAIL_REGEX = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")


class TemplateValidationError(Exception):
    """Raised when a template workbook is invalid."""


def read_template(
    template_path: Path | str, expected_field_names: Sequence[str]
) -> TemplateReadResult:
    """Read the Excel workbook and return normalized testcases."""
    path = Path(template_path)
    if not path.exists():
        raise TemplateValidationError(f"Template file not found: {path}")

    workbook = load_workbook(path, data_only=True)
    sheet = (
        workbook[TEMPLATE_SHEET_NAME]
        if TEMPLATE_SHEET_NAME in workbook.sheetnames
        else workbook.active
    )
    if sheet is None:
        raise TemplateValidationError("Template workbook has no active sheet.")
    assert isinstance(sheet, Worksheet)

    _validate_group_headers(sheet, len(expected_field_names))
    expected_columns = list(METADATA_COLUMNS + INPUT_COLUMNS + tuple(expected_field_names))
    header_values = [
        sheet.cell(row=2, column=index + 1).value for index in range(len(expected_columns))
    ]
    if header_values != expected_columns:
        raise TemplateValidationError("Template columns do not match the configured schema.")
    _ensure_no_extra_columns(sheet, len(expected_columns))

    header_map = {name: idx + 1 for idx, name in enumerate(expected_columns)}
    testcases = _parse_rows(sheet, header_map, expected_field_names)

    return TemplateReadResult(testcases=tuple(testcases))


def _validate_group_headers(sheet, expected_field_count: int) -> None:
    metadata_label = sheet.cell(row=1, column=1).value
    input_label = sheet.cell(row=1, column=len(METADATA_COLUMNS) + 1).value
    expected_label = sheet.cell(row=1, column=len(METADATA_COLUMNS) + len(INPUT_COLUMNS) + 1).value
    if metadata_label != "Metadata" or input_label != "Input" or expected_label != "Expected":
        raise TemplateValidationError("Template missing required group headers.")
    if expected_field_count == 0:
        raise TemplateValidationError("Expected fields list must not be empty.")


def _ensure_no_extra_columns(sheet, expected_count: int) -> None:
    for column in range(expected_count + 1, sheet.max_column + 1):
        value = sheet.cell(row=2, column=column).value
        if value not in (None, ""):
            raise TemplateValidationError("Template contains unexpected additional columns.")


def _parse_rows(
    sheet, header_map: Mapping[str, int], expected_field_names: Sequence[str]
) -> list[TemplateTestCase]:
    testcases: list[TemplateTestCase] = []
    seen_ids: set[str] = set()
    seen_pairs: dict[tuple[str, str], int] = {}
    for row_idx in range(3, sheet.max_row + 1):
        row_data = {
            name: sheet.cell(row=row_idx, column=col_index).value
            for name, col_index in header_map.items()
        }
        if _row_is_empty(row_data):
            continue
        testcase = _build_testcase(row_idx, row_data, expected_field_names)
        if testcase.test_id in seen_ids:
            raise TemplateValidationError(
                f"Duplicate ID '{testcase.test_id}' detected (row {row_idx})."
            )
        seen_ids.add(testcase.test_id)
        if testcase.enabled:
            pair = (testcase.from_address.lower(), testcase.subject.strip())
            previous = seen_pairs.get(pair)
            if previous:
                details = (
                    "Duplicate FROM/SUBJECT combination detected for rows "
                    f"{previous} and {row_idx}."
                )
                raise TemplateValidationError(details)
            seen_pairs[pair] = row_idx
        testcases.append(testcase)
    if not testcases:
        raise TemplateValidationError("Template does not contain any testcase rows.")
    return testcases


def _row_is_empty(row_data: Mapping[str, object]) -> bool:
    return all(_is_empty(value) for value in row_data.values())


def _build_testcase(
    row_number: int, row_data: Mapping[str, object], expected_field_names: Sequence[str]
) -> TemplateTestCase:
    test_id = _require_text(row_data["ID"], "ID", row_number)
    tags = _parse_tags(row_data.get("Tags"))
    enabled = _parse_bool(row_data.get("Enabled"))
    notes = _optional_string(row_data.get("Notes"))
    from_address = _require_text(row_data["FROM"], "FROM", row_number)
    if not EMAIL_REGEX.fullmatch(from_address):
        raise TemplateValidationError(f"Row {row_number}: invalid FROM address '{from_address}'.")
    subject = _require_text(row_data["SUBJECT"], "SUBJECT", row_number)
    body = _optional_string(row_data.get("BODY"))
    attachment = _optional_string(row_data.get("ATTACHMENT"))

    expected_values = {name: row_data.get(name) for name in expected_field_names}

    return TemplateTestCase(
        row_number=row_number,
        test_id=test_id,
        tags=tags,
        enabled=enabled,
        notes=notes,
        from_address=from_address,
        subject=subject,
        body=body,
        attachment=attachment,
        expected_values=expected_values,
    )


def _parse_tags(value: object) -> tuple[str, ...]:
    if _is_empty(value):
        return ()
    if not isinstance(value, str):
        value = str(value)
    tags = tuple(filter(None, (item.strip() for item in value.split(","))))
    return tags


def _parse_bool(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"", "true", "1", "yes"}:
            return True
        if lowered in {"false", "0", "no"}:
            return False
    if isinstance(value, int | float):
        return bool(value)
    raise TemplateValidationError(f"Unable to interpret boolean value: {value!r}")


def _optional_string(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty(value: object) -> bool:
    return _optional_string(value) == ""


def _require_text(value: object, column_name: str, row_number: int) -> str:
    if _is_empty(value):
        raise TemplateValidationError(f"Row {row_number}: column '{column_name}' is required.")
    return str(value).strip()
