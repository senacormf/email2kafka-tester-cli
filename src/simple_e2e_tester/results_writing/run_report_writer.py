"""Results workbook writer service."""

from __future__ import annotations

import hashlib
import json
from collections.abc import Mapping, Sequence
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from simple_e2e_tester.configuration.runtime_settings import SchemaConfig
from simple_e2e_tester.email_sending.delivery_outcomes import SendStatus
from simple_e2e_tester.matching_validation.matching_outcomes import (
    FieldMismatch,
    MatchValidationResult,
)
from simple_e2e_tester.schema_management.schema_models import FlattenedField
from simple_e2e_tester.template_generation import (
    INPUT_COLUMNS,
    METADATA_COLUMNS,
    SCHEMA_SHEET_NAME,
    TEMPLATE_SHEET_NAME,
)
from simple_e2e_tester.template_ingestion.testcase_models import TemplateTestCase

from .report_models import MatchStatus, RunMetadata


@dataclass(frozen=True)
class _SheetLayout:
    """Column layout for output sheet rendering."""

    expected_field_names: tuple[str, ...]
    input_columns: tuple[str, ...]
    original_column_count: int
    actual_start_column: int
    match_column: int


@dataclass(frozen=True)
class _RunCounts:
    """Computed run-level counters for the RunInfo sheet."""

    enabled_count: int
    matched: int
    passed: int
    failed: int
    not_found: int
    conflicts: int


@dataclass(frozen=True)
class _RowWriteContext:
    """Context needed while writing testcase rows."""

    layout: _SheetLayout
    matches_by_test_id: Mapping[str, Sequence]
    conflict_test_ids: set[str]
    unmatched_test_ids: set[str]
    send_status: Mapping[str, SendStatus]


# pylint: disable=too-many-arguments
def write_results_workbook(
    template_path: Path | str,
    output_path: Path | str,
    schema_config: SchemaConfig,
    schema_fields: Sequence[FlattenedField],
    testcases: Sequence[TemplateTestCase],
    match_result: MatchValidationResult,
    run_metadata: RunMetadata,
    send_status_by_test_id: Mapping[str, SendStatus] | None = None,
) -> None:
    """Write run output workbook with Actual and Match columns."""
    send_status = dict(send_status_by_test_id or {})
    layout = _build_sheet_layout(schema_fields)
    row_context = _build_row_write_context(
        layout=layout,
        match_result=match_result,
        send_status=send_status,
    )

    workbook = load_workbook(Path(template_path))
    sheet = (
        workbook[TEMPLATE_SHEET_NAME]
        if TEMPLATE_SHEET_NAME in workbook.sheetnames
        else workbook.active
    )
    _ensure_header_prefix(sheet, layout.input_columns)
    _write_actual_and_match_headers(
        sheet,
        layout.actual_start_column,
        layout.match_column,
        layout.expected_field_names,
    )

    _write_testcase_rows(
        sheet=sheet,
        testcases=testcases,
        row_context=row_context,
    )

    _write_schema_sheet(workbook, schema_config)
    _write_run_info_sheet(
        workbook=workbook,
        run_metadata=run_metadata,
        testcases=testcases,
        match_result=match_result,
        send_status=send_status,
    )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


# pylint: enable=too-many-arguments


def _build_sheet_layout(schema_fields: Sequence[FlattenedField]) -> _SheetLayout:
    expected_field_names = tuple(field.path for field in schema_fields)
    input_columns = tuple(METADATA_COLUMNS + INPUT_COLUMNS + expected_field_names)
    original_column_count = len(input_columns)
    actual_start_column = original_column_count + 1
    match_column = actual_start_column + len(expected_field_names)
    return _SheetLayout(
        expected_field_names=expected_field_names,
        input_columns=input_columns,
        original_column_count=original_column_count,
        actual_start_column=actual_start_column,
        match_column=match_column,
    )


def _build_row_write_context(
    *,
    layout: _SheetLayout,
    match_result: MatchValidationResult,
    send_status: Mapping[str, SendStatus],
) -> _RowWriteContext:
    conflict_test_ids = {
        test_id
        for conflict in match_result.conflicts
        for test_id in conflict.candidate_expected_event_ids
    }
    return _RowWriteContext(
        layout=layout,
        matches_by_test_id=_group_matches_by_test_id(match_result),
        conflict_test_ids=conflict_test_ids,
        unmatched_test_ids=set(match_result.unmatched_expected_event_ids),
        send_status=send_status,
    )


def _write_testcase_rows(
    *,
    sheet,
    testcases: Sequence[TemplateTestCase],
    row_context: _RowWriteContext,
) -> None:
    rows_added = 0
    for testcase in sorted(testcases, key=lambda case: case.row_number):
        rows_added += _write_single_testcase_rows(
            sheet=sheet,
            testcase=testcase,
            rows_added=rows_added,
            testcase_matches=row_context.matches_by_test_id.get(testcase.test_id, []),
            row_context=row_context,
        )


def _write_single_testcase_rows(
    *,
    sheet,
    testcase: TemplateTestCase,
    rows_added: int,
    testcase_matches: Sequence,
    row_context: _RowWriteContext,
) -> int:
    base_row = testcase.row_number + rows_added
    inserted_rows = _insert_duplicate_rows(
        sheet=sheet,
        base_row=base_row,
        match_count=len(testcase_matches),
        column_count=row_context.layout.original_column_count,
    )

    if testcase_matches:
        _write_match_rows(
            sheet=sheet,
            base_row=base_row,
            matches=testcase_matches,
            layout=row_context.layout,
        )
        return inserted_rows

    status = _resolve_unmatched_status(
        testcase=testcase,
        send_status=row_context.send_status.get(testcase.test_id),
        conflict_test_ids=row_context.conflict_test_ids,
        unmatched_test_ids=row_context.unmatched_test_ids,
    )
    sheet.cell(row=base_row, column=row_context.layout.match_column, value=status.value)
    return inserted_rows


def _insert_duplicate_rows(*, sheet, base_row: int, match_count: int, column_count: int) -> int:
    inserted_rows = 0
    if match_count <= 1:
        return inserted_rows
    for duplicate_index in range(1, match_count):
        new_row = base_row + duplicate_index
        sheet.insert_rows(new_row)
        _copy_row(
            sheet,
            source_row=base_row,
            target_row=new_row,
            column_count=column_count,
        )
        inserted_rows += 1
    return inserted_rows


def _write_match_rows(
    *,
    sheet,
    base_row: int,
    matches: Sequence,
    layout: _SheetLayout,
) -> None:
    for index, validated_match in enumerate(matches):
        row_number = base_row + index
        _write_actual_values(
            sheet,
            row_number=row_number,
            start_column=layout.actual_start_column,
            expected_field_names=layout.expected_field_names,
            flattened_values=validated_match.observed_event.flattened,
        )
        sheet.cell(
            row=row_number,
            column=layout.match_column,
            value=_format_mismatches(validated_match.mismatches),
        )


def _ensure_header_prefix(sheet, expected_columns: Sequence[str]) -> None:
    header_values = [
        sheet.cell(row=2, column=index).value for index in range(1, len(expected_columns) + 1)
    ]
    if header_values != list(expected_columns):
        raise ValueError("Template columns do not match expected schema-derived columns.")


def _write_actual_and_match_headers(
    sheet,
    actual_start_column: int,
    match_column: int,
    expected_field_names: Sequence[str],
) -> None:
    actual_end_column = match_column - 1
    if expected_field_names:
        _merge_and_label(sheet, "Actual", actual_start_column, actual_end_column)
    sheet.cell(row=1, column=match_column, value="Match")
    sheet.cell(row=1, column=match_column).style = "Headline 1"

    for index, field_name in enumerate(expected_field_names):
        column_index = actual_start_column + index
        sheet.cell(row=2, column=column_index, value=field_name)
        sheet.column_dimensions[get_column_letter(column_index)].width = max(
            12, min(len(field_name) + 6, 40)
        )
    sheet.cell(row=2, column=match_column, value="Match")
    sheet.column_dimensions[get_column_letter(match_column)].width = 50


def _merge_and_label(sheet, label: str, start_column: int, end_column: int) -> None:
    start_letter = get_column_letter(start_column)
    end_letter = get_column_letter(end_column)
    sheet.merge_cells(f"{start_letter}1:{end_letter}1")
    sheet.cell(row=1, column=start_column, value=label)
    sheet.cell(row=1, column=start_column).style = "Headline 1"


def _group_matches_by_test_id(match_result: MatchValidationResult) -> dict[str, list]:
    grouped: dict[str, list] = {}
    for validated_match in match_result.matches:
        grouped.setdefault(
            validated_match.expected_event.expected_event_id,
            [],
        ).append(validated_match)
    return grouped


def _copy_row(sheet, source_row: int, target_row: int, column_count: int) -> None:
    if source_row in sheet.row_dimensions:
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height

    for column in range(1, column_count + 1):
        source_cell = sheet.cell(row=source_row, column=column)
        target_cell = sheet.cell(row=target_row, column=column)
        target_cell.value = source_cell.value
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.fill = copy(source_cell.fill)
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)


def _write_actual_values(
    sheet,
    row_number: int,
    start_column: int,
    expected_field_names: Sequence[str],
    flattened_values: Mapping[str, Any],
) -> None:
    for index, field_name in enumerate(expected_field_names):
        column = start_column + index
        value = flattened_values.get(field_name)
        sheet.cell(row=row_number, column=column, value=_normalize_output_value(value))


def _normalize_output_value(value: Any) -> Any:
    if isinstance(value, Mapping):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    if isinstance(value, Sequence) and not isinstance(value, str | bytes):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def _format_mismatches(mismatches: Sequence[FieldMismatch]) -> str:
    if not mismatches:
        return MatchStatus.OK.value
    blocks = [
        f"expected: {mismatch.expected}\nactual: {mismatch.actual}" for mismatch in mismatches
    ]
    return "\n".join(blocks)


def _resolve_unmatched_status(
    testcase: TemplateTestCase,
    send_status: SendStatus | None,
    conflict_test_ids: set[str],
    unmatched_test_ids: set[str],
) -> MatchStatus:
    if send_status == SendStatus.FAILED:
        return MatchStatus.SEND_FAILED
    if send_status == SendStatus.SKIPPED or not testcase.enabled:
        return MatchStatus.SKIPPED
    if testcase.test_id in conflict_test_ids:
        return MatchStatus.CONFLICT
    if testcase.test_id in unmatched_test_ids:
        return MatchStatus.NOT_FOUND
    return MatchStatus.NOT_FOUND


def _write_schema_sheet(workbook, schema_config: SchemaConfig) -> None:
    if SCHEMA_SHEET_NAME in workbook.sheetnames:
        workbook.remove(workbook[SCHEMA_SHEET_NAME])
    sheet = workbook.create_sheet(SCHEMA_SHEET_NAME)
    schema_hash = hashlib.sha256(schema_config.text.encode("utf-8")).hexdigest()
    entries = (
        ("schema_type", schema_config.schema_type),
        ("schema_hash", schema_hash),
        ("schema_text", schema_config.text),
    )
    for row, (key, value) in enumerate(entries, start=1):
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=value)


def _write_run_info_sheet(
    workbook,
    run_metadata: RunMetadata,
    testcases: Sequence[TemplateTestCase],
    match_result: MatchValidationResult,
    send_status: Mapping[str, SendStatus],
) -> None:
    if "RunInfo" in workbook.sheetnames:
        workbook.remove(workbook["RunInfo"])
    sheet = workbook.create_sheet("RunInfo")
    counts = _calculate_run_counts(testcases, match_result, send_status)

    entries = (
        ("run_start", run_metadata.run_start.isoformat()),
        ("input_path", str(run_metadata.input_path)),
        ("output_path", str(run_metadata.output_path)),
        ("kafka_topic", run_metadata.kafka_topic),
        ("timeout_seconds", run_metadata.timeout_seconds),
        ("total", len(testcases)),
        ("enabled", counts.enabled_count),
        ("sent_ok", run_metadata.sent_ok),
        ("matched", counts.matched),
        ("passed", counts.passed),
        ("failed", counts.failed),
        ("not_found", counts.not_found),
        ("conflicts", counts.conflicts),
    )
    for row, (key, value) in enumerate(entries, start=1):
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=value)


def _calculate_run_counts(
    testcases: Sequence[TemplateTestCase],
    match_result: MatchValidationResult,
    send_status: Mapping[str, SendStatus],
) -> _RunCounts:
    enabled_count = sum(1 for testcase in testcases if testcase.enabled)
    matched = len(match_result.matches)
    passed = sum(1 for item in match_result.matches if not item.mismatches)
    failed_validation = sum(1 for item in match_result.matches if item.mismatches)
    send_failed = sum(1 for status in send_status.values() if status == SendStatus.FAILED)
    conflicts = len(match_result.conflicts)
    skipped_ids = {
        test_id for test_id, status in send_status.items() if status == SendStatus.SKIPPED
    }
    not_found_ids = set(match_result.unmatched_expected_event_ids) - {
        test_id
        for conflict in match_result.conflicts
        for test_id in conflict.candidate_expected_event_ids
    }
    not_found_ids -= skipped_ids
    return _RunCounts(
        enabled_count=enabled_count,
        matched=matched,
        passed=passed,
        failed=failed_validation + send_failed,
        not_found=len(not_found_ids),
        conflicts=conflicts,
    )
