"""Excel template generation service."""

from __future__ import annotations

import hashlib
from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from simple_e2e_tester.configuration.runtime_settings import SchemaConfig
from simple_e2e_tester.schema_management.schema_models import FlattenedField

from .constants import INPUT_COLUMNS, METADATA_COLUMNS, SCHEMA_SHEET_NAME, TEMPLATE_SHEET_NAME


def generate_template_workbook(
    schema_config: SchemaConfig,
    fields: Sequence[FlattenedField],
    output_path: Path | str,
) -> None:
    """Create the Excel template containing metadata, input, and expected columns."""
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("Workbook active sheet is not available.")
    assert isinstance(sheet, Worksheet)
    sheet.title = TEMPLATE_SHEET_NAME

    expected_columns = [field.path for field in fields]
    all_columns = list(METADATA_COLUMNS + INPUT_COLUMNS + tuple(expected_columns))

    _write_group_headers(sheet, len(METADATA_COLUMNS), len(INPUT_COLUMNS), len(expected_columns))
    for column_index, name in enumerate(all_columns, start=1):
        sheet.cell(row=2, column=column_index, value=name)
        sheet.column_dimensions[get_column_letter(column_index)].width = max(
            12, min(len(name) + 6, 40)
        )

    _write_schema_sheet(workbook, schema_config)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_group_headers(sheet, metadata_count: int, input_count: int, expected_count: int) -> None:
    groups = [
        ("Metadata", 1, metadata_count),
        ("Input", metadata_count + 1, input_count),
        ("Expected", metadata_count + input_count + 1, expected_count),
    ]
    for label, start_column, count in groups:
        if count <= 0:
            continue
        end_column = start_column + count - 1
        start_letter = get_column_letter(start_column)
        end_letter = get_column_letter(end_column)
        sheet.merge_cells(f"{start_letter}1:{end_letter}1")
        sheet[f"{start_letter}1"].value = label
        sheet[f"{start_letter}1"].style = "Headline 1"


def _write_schema_sheet(workbook: Workbook, schema_config: SchemaConfig) -> None:
    sheet = workbook.create_sheet(SCHEMA_SHEET_NAME)
    schema_hash = hashlib.sha256(schema_config.text.encode("utf-8")).hexdigest()
    entries = [
        ("schema_type", schema_config.schema_type),
        ("schema_hash", schema_hash),
        ("schema_text", schema_config.text),
    ]
    for row_index, (key, value) in enumerate(entries, start=1):
        sheet.cell(row=row_index, column=1, value=key)
        sheet.cell(row=row_index, column=2, value=value)
